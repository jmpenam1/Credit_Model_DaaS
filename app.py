import io
from dataclasses import dataclass
import numpy as np
import numpy_financial as npf
import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

st.set_page_config(page_title="Cotizador DaaS", layout="wide")

DEFAULT_ITEMS = pd.DataFrame(
    [
        {"Tipo": "Laptop", "Nombre": "Laptop Perfil 1", "Cantidad": 1, "Costo_unit": 100.0, "Spare_unit": 0.0},
        {"Tipo": "Monitor", "Nombre": "Monitor", "Cantidad": 1, "Costo_unit": 50.0, "Spare_unit": 0.0},
        {"Tipo": "Servicios", "Nombre": "Servicios cotizados", "Cantidad": 1, "Costo_unit": 50.0, "Spare_unit": 0.0},
    ]
)

def _as_float(x, default=0.0):
    try:
        if x is None or (isinstance(x, str) and x.strip() == ""):
            return float(default)
        return float(x)
    except Exception:
        return float(default)

def _as_int(x, default=0):
    try:
        if x is None or (isinstance(x, str) and x.strip() == ""):
            return int(default)
        return int(float(x))
    except Exception:
        return int(default)

@dataclass
class Params:
    plazo_meses: int
    margen: float
    iva_venta: float
    residual_pct: float
    tasa_coloc: float
    tasa_capt: float
    mantenimiento_pct: float
    seguros_pct: float
    provision_pct: float
    ica_pct: float
    renta_pct: float
    descuento: float

def compute_quote(items: pd.DataFrame, p: Params):
    df = items.copy()
    df["Cantidad"] = df["Cantidad"].apply(lambda v: max(_as_int(v, 0), 0))
    df["Costo_unit"] = df["Costo_unit"].apply(lambda v: max(_as_float(v, 0.0), 0.0))
    df["Spare_unit"] = df["Spare_unit"].apply(lambda v: max(_as_float(v, 0.0), 0.0))

    df["Costo_total"] = (df["Costo_unit"] + df["Spare_unit"]) * df["Cantidad"]

    denom = max(1e-9, (1.0 - p.margen))
    df["Venta_sin_IVA_unit"] = (df["Costo_unit"] + df["Spare_unit"]) / denom
    df["IVA_unit"] = df["Venta_sin_IVA_unit"] * p.iva_venta
    df["Venta_con_IVA_unit"] = df["Venta_sin_IVA_unit"] + df["IVA_unit"]
    df["Venta_total"] = df["Venta_con_IVA_unit"] * df["Cantidad"]

    total_costo = float(df["Costo_total"].sum())
    total_venta = float(df["Venta_total"].sum())

    n = int(max(p.plazo_meses, 1))

    pv_cli = total_venta
    fv_cli = -pv_cli * p.residual_pct
    canon_mensual = float(-npf.pmt(p.tasa_coloc, n, pv_cli, fv_cli)) if pv_cli > 0 else 0.0
    residual_cli = pv_cli * p.residual_pct if pv_cli > 0 else 0.0

    pv_bank = total_costo
    fv_bank = -pv_bank * p.residual_pct
    costo_fondeo_mensual = float(-npf.pmt(p.tasa_capt, n, pv_bank, fv_bank)) if pv_bank > 0 else 0.0
    residual_bank = pv_bank * p.residual_pct if pv_bank > 0 else 0.0

    summary = {
        "Costo_equipos": total_costo,
        "Venta_total_con_IVA": total_venta,
        "Canon_mensual_cliente": canon_mensual,
        "Residual_cliente": residual_cli,
        "Pago_mensual_fondeo": costo_fondeo_mensual,
        "Residual_fondeo": residual_bank,
    }
    return df, summary

def compute_cashflows(summary: dict, p: Params) -> pd.DataFrame:
    n = int(max(p.plazo_meses, 1))

    canon = float(summary["Canon_mensual_cliente"])
    resid_cli = float(summary["Residual_cliente"])
    pago_fondeo = float(summary["Pago_mensual_fondeo"])
    resid_fondeo = float(summary["Residual_fondeo"])
    costo_equipos = float(summary["Costo_equipos"])

    months = list(range(0, n + 1))
    rows = []
    for m in months:
        if m == 0:
            rows.append({
                "Mes": 0,
                "Cobro_cliente": 0.0,
                "Cobro_residual": 0.0,
                "Pago_fondeo": 0.0,
                "Pago_residual": 0.0,
                "Spread": 0.0,
                "Op_costos": 0.0,
                "Provision": 0.0,
                "ICA": 0.0,
                "Utilidad_AI": 0.0,
                "Impuesto": 0.0,
                "Flujo_neto": -costo_equipos,
            })
            continue

        cobro = canon
        cobro_res = resid_cli if m == n else 0.0

        pago = pago_fondeo
        pago_res = resid_fondeo if m == n else 0.0

        spread = (cobro + cobro_res) - (pago + pago_res)

        op = (costo_equipos * (p.mantenimiento_pct + p.seguros_pct) / n) if costo_equipos > 0 else 0.0
        prov = max(0.0, spread) * p.provision_pct
        ica = max(0.0, (cobro + cobro_res)) * p.ica_pct

        utilidad_ai = spread - op - prov - ica
        impuesto = max(0.0, utilidad_ai) * p.renta_pct

        rows.append({
            "Mes": m,
            "Cobro_cliente": cobro,
            "Cobro_residual": cobro_res,
            "Pago_fondeo": pago,
            "Pago_residual": pago_res,
            "Spread": spread,
            "Op_costos": op,
            "Provision": prov,
            "ICA": ica,
            "Utilidad_AI": utilidad_ai,
            "Impuesto": impuesto,
            "Flujo_neto": utilidad_ai - impuesto,
        })

    return pd.DataFrame(rows)

def npv_monthly(rate: float, cashflows: np.ndarray) -> float:
    if len(cashflows) == 0:
        return 0.0
    return float(cashflows[0] + npf.npv(rate, cashflows[1:]))

def export_excel(items_calc: pd.DataFrame, cashflows: pd.DataFrame, params: Params, summary: dict) -> bytes:
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Quote"
    ws2 = wb.create_sheet("Cashflows")
    ws3 = wb.create_sheet("Parametros")

    for r in dataframe_to_rows(items_calc, index=False, header=True):
        ws1.append(r)

    for r in dataframe_to_rows(cashflows, index=False, header=True):
        ws2.append(r)

    ws3.append(["Parametro", "Valor"])
    for k, v in {
        "plazo_meses": params.plazo_meses,
        "margen": params.margen,
        "iva_venta": params.iva_venta,
        "residual_pct": params.residual_pct,
        "tasa_coloc_mensual": params.tasa_coloc,
        "tasa_capt_mensual": params.tasa_capt,
        "mantenimiento_pct": params.mantenimiento_pct,
        "seguros_pct": params.seguros_pct,
        "provision_pct": params.provision_pct,
        "ica_pct": params.ica_pct,
        "renta_pct": params.renta_pct,
        "tasa_descuento_mensual": params.descuento,
    }.items():
        ws3.append([k, float(v)])

    ws3.append([])
    ws3.append(["Resumen", ""])
    for k, v in summary.items():
        ws3.append([k, float(v)])

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()

st.title("Cotizador DaaS (Streamlit)")

with st.sidebar:
    st.header("Parámetros")
    plazo_meses = st.number_input("Plazo (meses)", min_value=1, max_value=120, value=36, step=1)

    colA, colB = st.columns(2)
    with colA:
        margen = st.number_input("Margen (ej: 0.05)", min_value=0.0, max_value=0.9, value=0.05, step=0.01, format="%.4f")
        residual_pct = st.number_input("Residual % (ej: 0.15)", min_value=0.0, max_value=0.9, value=0.15, step=0.01, format="%.4f")
        iva_venta = st.number_input("IVA venta (ej: 0.19)", min_value=0.0, max_value=0.5, value=0.19, step=0.01, format="%.4f")
    with colB:
        tasa_coloc = st.number_input("Tasa colocación mensual (ej: 0.026)", min_value=0.0, max_value=0.5, value=0.026, step=0.001, format="%.4f")
        tasa_capt = st.number_input("Tasa captación mensual (ej: 0.006)", min_value=0.0, max_value=0.5, value=0.006, step=0.001, format="%.4f")
        descuento = st.number_input("Tasa descuento mensual (NPV)", min_value=0.0, max_value=0.5, value=float(tasa_capt), step=0.001, format="%.4f")

    st.divider()
    st.subheader("Costos / riesgos")
    colC, colD = st.columns(2)
    with colC:
        mantenimiento_pct = st.number_input("Mantenimiento % (sobre costo equipos)", min_value=0.0, max_value=0.5, value=0.01, step=0.005, format="%.4f")
        seguros_pct = st.number_input("Seguros % (sobre costo equipos)", min_value=0.0, max_value=0.5, value=0.02, step=0.005, format="%.4f")
        provision_pct = st.number_input("Provisión % (sobre spread)", min_value=0.0, max_value=0.5, value=0.02, step=0.005, format="%.4f")
    with colD:
        ica_pct = st.number_input("ICA % (sobre cobro cliente)", min_value=0.0, max_value=0.2, value=0.01, step=0.001, format="%.4f")
        renta_pct = st.number_input("Impuesto renta %", min_value=0.0, max_value=0.8, value=0.35, step=0.01, format="%.4f")

    st.caption("Si tus tasas están en E.A., conviértelas a mensual: (1+EA)**(1/12)-1")

params = Params(
    plazo_meses=int(plazo_meses),
    margen=float(margen),
    iva_venta=float(iva_venta),
    residual_pct=float(residual_pct),
    tasa_coloc=float(tasa_coloc),
    tasa_capt=float(tasa_capt),
    mantenimiento_pct=float(mantenimiento_pct),
    seguros_pct=float(seguros_pct),
    provision_pct=float(provision_pct),
    ica_pct=float(ica_pct),
    renta_pct=float(renta_pct),
    descuento=float(descuento),
)

st.subheader("Ítems de la cotización")

if "items" not in st.session_state:
    st.session_state["items"] = DEFAULT_ITEMS.copy()

b1, b2, _ = st.columns([1, 1, 3])
with b1:
    if st.button("➕ Agregar fila"):
        st.session_state["items"] = pd.concat(
            [st.session_state["items"], pd.DataFrame([{"Tipo":"", "Nombre":"", "Cantidad":1, "Costo_unit":0.0, "Spare_unit":0.0}])],
            ignore_index=True
        )
with b2:
    if st.button("♻️ Reset"):
        st.session_state["items"] = DEFAULT_ITEMS.copy()

edited = st.data_editor(
    st.session_state["items"],
    num_rows="dynamic",
    use_container_width=True,
    hide_index=True,
    column_config={
        "Cantidad": st.column_config.NumberColumn(min_value=0, step=1),
        "Costo_unit": st.column_config.NumberColumn(min_value=0.0, step=1.0, format="%.2f"),
        "Spare_unit": st.column_config.NumberColumn(min_value=0.0, step=1.0, format="%.2f"),
    },
)
st.session_state["items"] = edited

items_calc, summary = compute_quote(st.session_state["items"], params)
cashflows = compute_cashflows(summary, params)

st.subheader("Indicadores")
c1, c2, c3, c4 = st.columns(4)
c1.metric("Costo equipos", f"${summary['Costo_equipos']:,.0f}")
c2.metric("Venta total (con IVA)", f"${summary['Venta_total_con_IVA']:,.0f}")
c3.metric("Canon mensual (cliente)", f"${summary['Canon_mensual_cliente']:,.0f}")
c4.metric("Pago mensual fondeo", f"${summary['Pago_mensual_fondeo']:,.0f}")

cf = cashflows["Flujo_neto"].to_numpy(dtype=float)
irr_m = float(npf.irr(cf)) if (np.any(cf != 0) and len(cf) >= 2) else 0.0
irr_ea = (1.0 + irr_m) ** 12 - 1.0 if irr_m > -1 else float("nan")
npv = npv_monthly(params.descuento, cf)

c5, c6, c7 = st.columns(3)
c5.metric("NPV (mensual)", f"${npv:,.0f}")
c6.metric("IRR mensual", f"{irr_m*100:.2f}%")
c7.metric("IRR E.A.", f"{irr_ea*100:.2f}%")

st.divider()

tab1, tab2, tab3 = st.tabs(["Detalle Quote", "Flujos de caja", "Descargar Excel"])

with tab1:
    st.dataframe(items_calc, use_container_width=True)

with tab2:
    st.dataframe(cashflows, use_container_width=True)
    st.line_chart(cashflows.set_index("Mes")["Flujo_neto"].cumsum())

with tab3:
    xlsx_bytes = export_excel(items_calc, cashflows, params, summary)
    st.download_button(
        "⬇️ Descargar Excel (Quote + Cashflows + Parametros)",
        data=xlsx_bytes,
        file_name="cotizacion_daas.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

st.caption("Modelo inspirado en tu template: margen para venta, PMT con residual, y flujo neto después de costos, provisión, ICA e impuestos.")
